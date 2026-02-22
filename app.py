import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# --- 樣式定義 (數據缺失改為黑底，數值異常改為粉/紅底) ---
STYLE_MISSING = {"fill": PatternFill("solid", fgColor="000000"), "font": Font(name="微軟正黑體", size=12, color="FFFFFF", bold=True)}
STYLE_LOW     = {"fill": PatternFill("solid", fgColor="FF0000"), "font": Font(name="微軟正黑體", size=12, color="FFFFFF", bold=True)}
STYLE_CAL     = {"fill": PatternFill("solid", fgColor="FFCCFF"), "font": Font(name="微軟正黑體", size=12, color="800000", bold=True)}

def clean_to_num(val):
    # 判斷是否為 NaN 或 空字串
    if pd.isna(val) or str(val).strip().lower() in ["", "nan", "none"]: 
        return None
    try:
        # 提取數字（處理可能包含單位的情況）
        res = re.findall(r"\d+\.?\d*", str(val))
        return float(res[0]) if res else 0.0
    except: 
        return 0.0

def alison_nuanhe_audit_v4(file):
    wb = load_workbook(file)
    sheets_df = pd.read_excel(file, sheet_name=None, header=None)
    logs = []
    total_scan = 0
    
    # 暖禾高標規範
    LIMITS = {"熱量": (750, 850), "全榖": 4.0, "豆魚": 4.0, "蔬菜": 2.0}

    for sn, df in sheets_df.items():
        ws = wb[sn]
        # 1. 尋找日期行 (鎖定 C 欄)
        d_row = None
        for i in range(min(20, len(df))):
            if any(k in str(df.iloc[i, 2]) for k in ["日期", "Date"]):
                d_row = i
                break
        if d_row is None: continue

        # 2. 橫向掃描週一到週五 (C-G 欄)
        for col in range(2, 7):
            if col >= len(df.columns): break
            
            date_cell = str(df.iloc[d_row, col]).strip()
            if "202" not in date_cell: continue
            this_day = date_cell.split(" ")[0]

            # 3. 縱向搜尋營養標籤 (限制日期下方 25 列，解決 228 項 nan 問題)
            for r_off in range(1, 26):
                r_idx = d_row + r_off
                if r_idx >= len(df): break
                
                # 模糊匹配：合併 B, C 欄標籤
                label = str(df.iloc[r_idx, 1]) + str(df.iloc[r_idx, 2])
                raw_val = df.iloc[r_idx, col]
                
                for key, limit in LIMITS.items():
                    if key in label:
                        total_scan += 1
                        val = clean_to_num(raw_val)
                        cell = ws.cell(row=r_idx+1, column=col+1)
                        
                        # --- 核心判定邏輯 ---
                        if val is None: # 狀況 A：完全沒填
                            cell.fill, cell.font = STYLE_MISSING["fill"], STYLE_MISSING["font"]
                            cell.value = "❌數據缺失"
                            logs.append({"分頁": sn, "日期": this_day, "項目": label, "原因": "數據缺失"})
                            
                        elif key == "熱量": # 狀況 B：熱量範圍
                            if val < limit[0] or val > limit[1]:
                                cell.fill, cell.font = STYLE_CAL["fill"], STYLE_CAL["font"]
                                logs.append({"分頁": sn, "日期": this_day, "項目": label, "原因": f"數值異常({val})"})
                                
                        else: # 狀況 C：營養份數 (0.0 安全過關，(0, limit) 報警)
                            if 0 < val < limit:
                                cell.fill, cell.font = STYLE_LOW["fill"], STYLE_LOW["font"]
                                logs.append({"分頁": sn, "日期": this_day, "項目": label, "原因": f"份數不足({val})"})
    
    out = BytesIO()
    wb.save(out)
    return logs, out.getvalue(), total_scan

# --- UI 介面 ---
st.title("🛡️ 輕食區(暖禾) 菜單自主稽核系統")
st.caption("製作者：Alison")

up = st.file_uploader("📂 上傳暖禾菜單 Excel", type=["xlsx"])
if up:
    try:
        results, excel_data, count = alison_nuanhe_audit_v4(up)
        st.info(f"📊 確實度：本次共精確稽核了 {count} 個營養數據點。")
        
        if count < 10:
            st.error("❌ 格式不對！程式找不到關鍵標籤，請確認 Excel 是否為暖禾格式。")
        elif results:
            st.error(f"🚩 發現 {len(results)} 處數據缺失或不符規範。")
            st.table(pd.DataFrame(results))
            st.download_button("📥 下載標註檔", excel_data, f"退件_{up.name}")
        else:
            st.success("🎉 完美！所有數據皆填寫完整且符合規範（0.0 已驗證有效）。")
    except Exception as e:
        st.error(f"🔥 系統崩潰：{str(e)}。")
